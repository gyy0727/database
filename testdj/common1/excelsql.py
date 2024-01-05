import json

import pymysql
import xlrd
import xlwt
import openpyxl

connection = pymysql.connect(
    host='localhost',  # 数据库地址
    user='root',  # 数据库用户名
    password='12345678',  # 数据库密码
    db='volunteeradmissionsystem',  # 数据库名称
    # charset = 'utf8 -- UTF-8 Unicode'
)
cursor = connection.cursor()
workbook = openpyxl.load_workbook(r"D:\computerman\testdj\平行志愿测试数据.xlsx")
# 选择工作表
sheet1 = workbook.active
sheet2 = workbook["招生计划表"]


# for column in sheet.columns:
#     for cell in column:
#         print(cell.value)
# for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
#     for cell in row:
#         if(cell.value=="KSH考生号"):
#             continue
#         value=str(cell.value)
#         sql="INSERT INTO students (考生号) VALUES ("+value+");"
#         cursor.execute(sql)
#         result = cursor.fetchall()
#         print(type(result),cursor.rowcount)
#         print(result)
#         print(cell.value)
# 关闭工作簿
# sql="INSERT INTO students (考生号,姓名,专业组号,投档成绩,服从调剂,专业志愿1,专业志愿2,专业志愿3,专业志愿4,专业志愿5,专业志愿6,外语语种,体检筛选备注) VALUES (1,2,3,4,5,6,7,8,9,10,11,12,13);"
#
workbook.close()
def selectStudent():
    sql = "select * from students"
    cursor.execute(sql)
    result = cursor.fetchall()
    return result
def selectMajor():
    sql = "select * from major"
    cursor.execute(sql)
    result = cursor.fetchall()
    return result
def selectAdmission():
    sql = "select * from admission"
    cursor.execute(sql)
    result = cursor.fetchall()
    return result
def selectWithdrawal():
    sql = "select * from withdrawal"
    cursor.execute(sql)
    result = cursor.fetchall()
    return result
def selectAdjustment():
    sql = "select * from adjustment"
    cursor.execute(sql)
    result = cursor.fetchall()
    return result
def selectAdmittedMajor(ZYDH):
    sql = "select * from major where 专业代号 ="+ZYDH
    cursor.execute(sql)
    result = cursor.fetchall()
    return result

def insertAddmitted():
    sql = """
    UPDATE admitted
    JOIN major ON admitted.专业代号 = major.专业代号
    SET admitted.待录取人数 = major.招生计划数;

            """
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result



def createTable():
    sql = '''   
CREATE TABLE withdrawal (
    退档位次 INT AUTO_INCREMENT PRIMARY KEY,
    考生号 VARCHAR(255),
    姓名 VARCHAR(255),
    FOREIGN KEY (考生号) REFERENCES students(考生号)
);
'''
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result


def orderStudent():
    sql='''
    SELECT * FROM students
        ORDER BY 投档成绩 DESC;
    '''
    cursor.execute(sql)
    result = cursor.fetchall()
    # result =json.dumps(result,ensure_ascii=False)
    return result

def selectMajorAdmitted():
    sql='''
    SELECT * FROM admitted WHERE 待录取人数 != 0;
    '''
    cursor.execute(sql)
    result = cursor.fetchall()
    # result =json.dumps(result,ensure_ascii=False)
    return result
def orderStudentJson():
    sql='''
    SELECT * FROM students
        ORDER BY 投档成绩 DESC;
    '''
    cursor.execute(sql)
    result = cursor.fetchall()
    result =json.dumps(result,ensure_ascii=False)
    return result
def withdrawaling(admitted,withdrawal,studentDetail):
    if not admitted:
        print("退档")
        sql4 = "insert into withdrawal (退档位次,考生号,姓名) values (%s,%s,%s)"
        update(sql4, (str(withdrawal), studentDetail[0][0], studentDetail[0][1]))

def orderAdjustment():
    sql='''
    SELECT * FROM adjustment
        ORDER BY 调剂位次 ASC;
    '''
    cursor.execute(sql)
    result = cursor.fetchall()
    # result =json.dumps(result,ensure_ascii=False)
    return result
def JudgingStrings(a,b):
    text = a
    char_to_find = b
    if text.find(char_to_find) != -1:
        return True
    else:
        return False


def insert(sql,data):
    cursor = connection.cursor()
    # 拼接并执行SQL语句
    cursor.executemany(sql, data)
    # 涉及写操作要提交
    connection.commit()

def orderJson():
    sql="""SELECT * FROM admission ORDER BY 学院名称,专业代号,录取位次;"""
    cursor.execute(sql)
    result = cursor.fetchall()
    # result=list(result)
    result =json.dumps(result,ensure_ascii=False)
    print(result)
    return result

def selectMajorOnly(major):
    sql = "select * from major where 专业名称=%s"
    cursor.execute(sql, (major,))
    result = cursor.fetchall()
    return result

def selectStudentOnly(student):
    sql = "select * from students where 考生号=%s"
    cursor.execute(sql, (student,))
    result = cursor.fetchall()
    return result

def selectAdmitted(ZYDH):
    sql = 'SELECT * FROM admitted where 专业代号='+ZYDH
    cursor.execute(sql)
    result = cursor.fetchall()
    return result

def update(sql,data):
    # 执行SQL语句
    cursor.execute(sql,data)
    connection.commit()
    result = cursor.fetchall()
    return result



def delete1():
    sql="DELETE FROM admission;"
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result
def delete2():
    sql="DELETE FROM adjustment;"
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result

def delete3():
    sql="DELETE FROM withdrawal;"
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result



def admitteing(student,major_,order):
    sql1 = "insert into admission (录取位次,考生号,姓名,专业组号,专业代号,专业名称,学院名称) values (%s,%s,%s,%s,%s,%s,%s)"
    update(sql1, (str(order), student[0], student[1], major_[0], major_[1], major_[2], major_[3]))
    sql2 = """
    UPDATE admitted 
    SET 待录取人数 = 待录取人数 - 1 
    WHERE 专业代号 = %s;
    """
    update(sql2, (major_[1],))

def admitteingAdjustment(student,major_,order):
    sql1 = "insert into admission (录取位次,考生号,姓名,专业组号,专业代号,专业名称,学院名称) values (%s,%s,%s,%s,%s,%s,%s)"
    update(sql1, (str(order), student[0][0], student[0][1], major_[0][0], major_[0][1], major_[0][2], major_[0][3]))
    sql2 = """
    UPDATE admitted 
    SET 待录取人数 = 待录取人数 - 1 
    WHERE 专业代号 = %s;
    """
    update(sql2, (major_[0][1],))

def test():
    sql="""
    UPDATE admitted 
    SET 待录取人数 = 待录取人数 - 1 
    WHERE 专业代号 = '021';"""
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result

def selectDepartmentDetial():
    sql="SELECT TABLE_NAME AS view_name FROM INFORMATION_SCHEMA.VIEWS WHERE TABLE_SCHEMA = 'volunteeradmissionsystem'; "
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result
def createView():
    sql="""
    CREATE VIEW admission_with_scores AS
    SELECT admission.*, students.成绩
    FROM admission
    JOIN students ON admission.考生号 = students.考生号
    ORDER BY 学院名称, 专业代号, 录取位次;"""
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result
def selectView(sql):
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result
def createViewEvery():
    sql="SELECT DISTINCT 学院名称 FROM admission;"
    cursor.execute(sql)
    connection.commit()
    result1 = cursor.fetchall()
    for result in result1:
        sql=" CREATE VIEW "+result[0]+" AS SELECT * FROM admission_with_scores WHERE 学院名称 = %s;"

        data=[result[0],result[0]]

        update(sql,result[0])
        print(result[0])


def getAVGScore(department):
    sql="SELECT AVG(投档成绩) AS average_score FROM "+department+";"
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result

def getMAX(department):
    sql="SELECT * FROM "+department+" WHERE `录取位次` = (SELECT MAX(录取位次) FROM "+department+");"
    cursor.execute(sql)
    connection.commit()
    result= cursor.fetchall()
    return result
def getMin(department):
    sql="SELECT * FROM "+department+" WHERE `录取位次` = (SELECT Min(录取位次) FROM "+department+");"
    cursor.execute(sql)
    connection.commit()
    result = cursor.fetchall()
    return result

def main():
    # delete1()
    # delete2()
    # delete3()
    # insertAddmitted()
    # order()
    # orderJson()
    createViewEvery()
if __name__ == "__main__":
    main()
