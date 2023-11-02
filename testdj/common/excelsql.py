import pymysql
import xlrd
import xlwt
import openpyxl

# 打开XLSX文件
workbook = openpyxl.load_workbook(r"D:\computerman\testdj\平行志愿测试数据.xlsx")

# 选择工作表
sheet = workbook.active

# 从工作表中读取数据
for row in sheet.iter_rows():
    for cell in row:
        print(cell.value)

# 关闭工作簿
workbook.close()

#建立一个MySQL连接
conn = pymysql.connect(
        host='localhost',
        user='root',
        passwd='12345678',
        db='mysql',
        port=3306,
        charset='utf8'
        )
# 获得游标
# cur = conn.cursor()
# # 创建插入SQL语句
# query = 'insert into test0615 (学号,姓名,年龄,性别,班级,分数,排名) values (%s, %s, %s, %s, %s, %s, %s)'
# # 创建一个for循环迭代读取xls文件每行数据的, 从第二行开始是要跳过标题行
# for r in range(1, sheet.nrows):
#       学号 = sheet.cell(r,0).value
#       姓名 = sheet.cell(r,1).value
#       年龄 = sheet.cell(r, 2).value
#       性别 = sheet.cell(r, 3).value
#       班级 = sheet.cell(r, 4).value
#       分数 = sheet.cell(r, 5).value
#       排名 = sheet.cell(r, 6).value
#       values = (学号,姓名,年龄,性别,班级,分数,排名)
#       # 执行sql语句
#       cur.execute(query, values)
# cur.close()
# conn.commit()
# conn.close()
# columns = str(sheet.ncols)
# rows = str(sheet.nrows)
# print ("导入 " +columns + " 列 " + rows + " 行数据到MySQL数据库!")
